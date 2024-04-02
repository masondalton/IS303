# Team: Duces
# Ethan - Tafi
# James - Santi
# Rex - Chips
# Kaitlyn - Concatakaitlyn
# Jack - Ninja
# Mason Dalton - Maestro Helper
# A program for organize data about subjects, students, and grades. Including subject statistics


#import all necessary libraries-- OS, subprocess for mac, 
#platform, workbook, get_column_letter, Font
import os
import subprocess
import platform
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Function for clearing terminal
def clear_screen():
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')

#Clears terminal screen after running
clear_screen()

# Importing the excel file we're using and activating worksheet for easy access
poorWorkbook = openpyxl.load_workbook("Poorly_Organized_Data_1.xlsx")

#Set active sheet with currSheet object
currSheet = poorWorkbook.active

# Create Workbook object for organizing the data, remove default sheet
organizedWorkbook = Workbook()
organizedWorkbook.remove(organizedWorkbook["Sheet"])

# Loop through poor data, creating organized sheets in workbook
for row in currSheet.iter_rows(min_row = 2, max_col = 3, max_row = currSheet.max_row):
    # Taking data about class subject, student name, id and grade, and store them into objects
    className = row[0].value
    studentInfo = row[1].value
    grade = row[2].value

    # Break first, last name, and id apart and assign to loop variables
    studentInfo = studentInfo.split("_")

    lastName = studentInfo[0]
    firstName = studentInfo[1]
    studID = studentInfo[2]

    # Select statement for creating new sheet if new subject or assigning existing Subject Sheet 
    if className not in organizedWorkbook.sheetnames :
        # Create a sheet in organized workbook based on new subject
        newSheet = organizedWorkbook.create_sheet(title=className)

        # As new sheet is create, these are the column headers
        newSheet.append(["Last Name", "First Name", "Student ID", "Grade"])

        # These makes a filter for all headers
        newSheet.auto_filter.ref = "a1:d1"
    else:
        # Assign existing subject sheet as new sheet to later append current row student variables
        newSheet = organizedWorkbook[className]

    # Take the new sheet or existing sheet and add the row variables
    newSheet.append([lastName, firstName, studID, grade])

# Loop for sumamrizing data of each subject
for sheet in organizedWorkbook.sheetnames :
    # Switch current sheet to the new, organized workbook based on place in loop
    currSheet = organizedWorkbook[sheet]

    # Max row and columns for positioning formulas
    iMaxRow = currSheet.max_row
    iMaxCol = currSheet.max_column + 1

    # Group of similar assignements to make summary table titles
    currSheet[get_column_letter(iMaxCol + 1) + "1"] = "Summary Statistics"
    currSheet[get_column_letter(iMaxCol + 1) + "2"] = "Highest Grade"
    currSheet[get_column_letter(iMaxCol + 1) + "3"] = "Lowest Grade"
    currSheet[get_column_letter(iMaxCol + 1) + "4"] = "Mean Grade"
    currSheet[get_column_letter(iMaxCol + 1) + "5"] = "Median Grade"
    currSheet[get_column_letter(iMaxCol + 1) + "6"] = "Number of Students"
      
    # Group of assignments for values of summary statistics
    currSheet[get_column_letter(iMaxCol + 2) + "1"] = "Value"
    currSheet[get_column_letter(iMaxCol + 2) + "2"] = "=MAX(D2:D" + str(iMaxRow) + ")"
    currSheet[get_column_letter(iMaxCol + 2) + "3"] = "=MIN(D2:D" + str(iMaxRow) + ")"
    currSheet[get_column_letter(iMaxCol + 2) + "4"] = "=AVERAGE(D2:D" + str(iMaxRow) + ")"
    currSheet[get_column_letter(iMaxCol + 2) + "5"] = "=MEDIAN(D2:D" + str(iMaxRow) + ")"
    currSheet[get_column_letter(iMaxCol + 2) + "6"] = "=COUNT(D2:D" + str(iMaxRow) + ")"

    # Styling loop for bold for headers in row 1 and column width being +5 more than length.
    for cell in currSheet["1:1"]:
        cell.font = Font(bold=True)
        if cell.value is not None:
            # add 5 to each column width
            currSheet.column_dimensions[cell.column_letter].width = len(str(cell.value)) + 5

# Name of new excel file
output_file = "formatted_grades_practice.xlsx"
# Save organized workbook to the file
organizedWorkbook.save(filename=output_file)
# Open the newly created file in excel
if platform.system() == 'Windows' :
    os.system(f'start excel {output_file}')
else : 
    # For opening file on mac
    subprocess.Popen(['open', output_file])

# Close out memory... for safety I guess
organizedWorkbook.close()